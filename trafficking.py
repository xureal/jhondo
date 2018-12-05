'''
This is the main automation file that does the grunt work. Placement names, creative names and URLs are read from a sheet (testtest) and are then assigned in the campaign specified in the input.
Before running the script make sure to upload the placements and creatives to the DCM campaign that is supposed to be trafficked.
Since the creative and placement lookup can only happen by names (labels), the names on the sheet need to EXACTLY match the creative names and the names of the placements uploaded to DCM.
You can use the creativecheck and placementcheck scripts to validate that they both do.
'''

import argparse
from oauth2client import file, client, tools
from apiclient.discovery import build
from httplib2 import Http
import config
import dictionaries
import sys
import time
import openpyxl
from datetime import datetime
from openpyxl.styles import Font

############stupid shit to get started##################

parser = argparse.ArgumentParser(parents=[tools.argparser])
flags = parser.parse_args()

CLIENT_SECRET = 'client_secret.json' # downloaded JSON file
SCOPES = ('https://www.googleapis.com/auth/dfareporting','https://www.googleapis.com/auth/dfatrafficking')

store = file.Storage('storage.json')
creds = store.get()
if not creds or creds.invalid:
    flow = client.flow_from_clientsecrets(CLIENT_SECRET, SCOPES)
    creds = tools.run_flow(flow, store, flags)

DFA = build('dfareporting', 'v3.2', http=creds.authorize(Http()))

print('Connected to Google servers')
##############end stupid shit##########################

#Input the values for advertiser name and campaign ID
advertiser_name = input('Please enter the advertiser name: ')
campaign_id = input('Please enter the campaign ID: ')
print('\n')
#Check if the entered advertiser is in the alias dict of the config file or the main advertiser dict, if not then exit the program
if advertiser_name in config.alias:
    advertiser_id = dictionaries.advertisers(config.alias[advertiser_name])[0]
    profile_id = dictionaries.advertisers(config.alias[advertiser_name])[2]
elif dictionaries.advertisers(advertiser_name) != 0:
    advertiser_id = dictionaries.advertisers(advertiser_name)[0]
    profile_id = dictionaries.advertisers(advertiser_name)[2]
else:
    print('Advertiser doesn\'t exist, please add the advertiser to the dictionary to proceed')
    print('Terminating...')
    sys.exit()

#Initizlize all of the required variables
start_time = time.time()
workbook = openpyxl.load_workbook(config.feed_path)
sheet = workbook.get_sheet_by_name(config.feed_sheet)
row_count = 0
skip_count = 0
today = datetime.today()


for i in range(1,sheet.max_row,1): #Find the total number of non-empty rows in the testtest sheet
    if sheet['A'+str(i)].value!=None and sheet['A'+str(i)].value!='':
        row_count += 1


for j in range(2,row_count+1,1): #Loop through all of the placements on the testtest sheet
    
    #Variables that need to be initialized per placement
    url_list = []
    creative_list = []
    exitid_list = []



    feed_placement_name = sheet['A'+str(j)].value #Get placement name from the feed
    print('Feed placement name - ' + feed_placement_name)

    if profile_id == dictionaries.profiles('Visa',config.user) or profile_id == dictionaries.profiles('Samsung',config.user):
        ad_name = sheet['H'+str(j)].value
    else:
        ad_name = feed_placement_name

    fetch_dcm_placement = DFA.placements().list(profileId=profile_id,campaignIds=campaign_id,advertiserIds=advertiser_id,searchString=feed_placement_name,archived=False).execute() #Fetch the placement from the DCM campaign by name
    dcm_placement_id = fetch_dcm_placement['placements'][0]['id'] #Get the placement ID of the fetched placement
    print('DCM placement ID - ' + str(dcm_placement_id))

    dcm_placement_startdate = today.strftime('%Y-%m-%d') #Setup the start date of the placement to today (for creative assignment purposes later)
    dcm_placement_enddate = fetch_dcm_placement['placements'][0]['pricingSchedule']['endDate'] #Get the placement end date based on the fetched placement
    
    fetch_dcm_campaign_enddate = DFA.campaigns().get(profileId=profile_id, id=campaign_id,fields='endDate').execute() #Fetch the campaign end date from DCM
    dcm_campaign_enddate = fetch_dcm_campaign_enddate['endDate'] #Get the campaign end date from the fetched data above
    
    try: #Try to find creatives and assign them
        for k in range(2,8,2): #Loop through all of the creatives on the testtest sheet for this placement
            feed_creative_name = sheet.cell(row=j,column=k).value
            if feed_creative_name: #Get creative name from the feed
                print('Feed creative name - ' + feed_creative_name)
                fetch_dcm_creative_list = DFA.creatives().list(profileId=profile_id,active=True,advertiserId=advertiser_id,campaignId=campaign_id,archived=False,searchString=feed_creative_name).execute() #Search for the creative name in the DCM campaign
                dcm_creative_id = fetch_dcm_creative_list['creatives'][0]['id'] #Get the ID of the creative(s) matching the search criteria

                fetch_dcm_creative = DFA.creatives().get(profileId=profile_id,id=dcm_creative_id).execute() #Fetch the creative asset by ID returned by the function above
                creative_list.append(dcm_creative_id) #Append the creative ID to a list that will be used for creative assignment later
                print('DCM Creative ID - ' + str(creative_list[int((k/2)-1)]))

                dcm_creative_type = fetch_dcm_creative['type'] #Get the type of the creative that was fetched by the function above
                print('The creative type is - ' +str(dcm_creative_type))

                if 'exitCustomEvents' in fetch_dcm_creative: #Check if the fetched creative has custom exit events
                    typeflag = True #Set the type flag to true if custom exits exists i.e. it's not a static banner
                    for l in range(0,len(fetch_dcm_creative['exitCustomEvents']),1): #Go through all of the custom exit events in the creative
                        dcm_creative_customevent = fetch_dcm_creative['exitCustomEvents'][l]['advertiserCustomEventName'] #Fetch the custom exit event in the creative
                        if ((type(dcm_creative_customevent) is str) and dcm_creative_customevent.lower() != 'backup'): #Make sure the exit event is not the backup one
                            dcm_creative_exitid = fetch_dcm_creative['exitCustomEvents'][l]['id'] #Fetch the id of the custom exit event
                            exitid_list.append(dcm_creative_exitid) #Append the exit event id to a list that will be used for creative assignment later
                            print('Creative exit name - ' + str(dcm_creative_customevent))
                else:
                    typeflag = False #Set the type flag to false if custom exits don't exist i.e. it's a static banner

            feed_url = sheet.cell(row=j,column=k+1).value #Get url from the feed
            if feed_url:
                url_list.append(feed_url) #Append the feed url to a list that will be used for creative assignment later
                print('Feed URL - ' + str(url_list[int((k/2)-1)]))
        if (len(creative_list)==1 and dcm_creative_type=='TRACKING_TEXT'):
            adscreate = DFA.ads().insert(profileId=profile_id,body={'campaignId':campaign_id,'advertiserId':advertiser_id,'name':ad_name,'startTime':dcm_placement_startdate+'T05:00:00Z','endTime':dcm_placement_enddate+'T00:00:00Z','type':'AD_SERVING_TRACKING','deliverySchedule':{'priority':'AD_PRIORITY_13','impressionRatio':1},'placementAssignments':[{'active':True,'placementId':dcm_placement_id}],'creativeRotation':{'creativeAssignments':[{'active':True,'creativeId':creative_list[0],'clickThroughUrl':{'customClickThroughUrl':url_list[0]}}]},'active':True}).execute()
        elif len(creative_list)==1 and typeflag == False:
            dcm_create_ad = DFA.ads().insert(profileId=profile_id,body={'campaignId':campaign_id,'advertiserId':advertiser_id,'name':ad_name,'startTime':dcm_placement_startdate+'T05:00:00Z','endTime':dcm_placement_enddate+'T00:00:00Z','type':'AD_SERVING_STANDARD_AD','deliverySchedule':{'priority':'AD_PRIORITY_13','impressionRatio':1},'placementAssignments':[{'active':True,'placementId':dcm_placement_id}],'creativeRotation':{'creativeAssignments':[{'active':True,'creativeId':creative_list[0],'clickThroughUrl':{'customClickThroughUrl':url_list[0]}}]},'active':True}).execute()
        elif len(creative_list)==2 and typeflag == False:
            dcm_create_ad = DFA.ads().insert(profileId=profile_id,body={'campaignId':campaign_id,'advertiserId':advertiser_id,'name':ad_name,'startTime':dcm_placement_startdate+'T05:00:00Z','endTime':dcm_placement_enddate+'T00:00:00Z','type':'AD_SERVING_STANDARD_AD','deliverySchedule':{'priority':'AD_PRIORITY_13','impressionRatio':1},'placementAssignments':[{'active':True,'placementId':dcm_placement_id}],'creativeRotation':{'creativeAssignments':[{'active':True,'creativeId':creative_list[0],'clickThroughUrl':{'customClickThroughUrl':url_list[0]}},{'active':True,'creativeId':creative_list[1],'clickThroughUrl':{'customClickThroughUrl':url_list[1]}}]},'active':True}).execute()
        elif len(creative_list)==3 and typeflag == False:
            dcm_create_ad = DFA.ads().insert(profileId=profile_id,body={'campaignId':campaign_id,'advertiserId':advertiser_id,'name':ad_name,'startTime':dcm_placement_startdate+'T05:00:00Z','endTime':dcm_placement_enddate+'T00:00:00Z','type':'AD_SERVING_STANDARD_AD','deliverySchedule':{'priority':'AD_PRIORITY_13','impressionRatio':1},'placementAssignments':[{'active':True,'placementId':dcm_placement_id}],'creativeRotation':{'creativeAssignments':[{'active':True,'creativeId':creative_list[0],'clickThroughUrl':{'customClickThroughUrl':url_list[0]}},{'active':True,'creativeId':creative_list[1],'clickThroughUrl':{'customClickThroughUrl':url_list[1]}},{'active':True,'creativeId':creative_list[2],'clickThroughUrl':{'customClickThroughUrl':url_list[2]}}]},'active':True}).execute()
        elif len(creative_list)==1: #If there is only one creative on rotation
            dcm_create_ad = DFA.ads().insert(profileId=profile_id,body={'campaignId':campaign_id,'advertiserId':advertiser_id,'name':ad_name,'startTime':dcm_placement_startdate+'T05:00:00Z','endTime':dcm_placement_enddate+'T00:00:00Z','type':'AD_SERVING_STANDARD_AD','deliverySchedule':{'priority':'AD_PRIORITY_13','impressionRatio':1},'placementAssignments':[{'active':True,'placementId':dcm_placement_id}],'creativeRotation':{'creativeAssignments':[{'active':True,'creativeId':creative_list[0],'clickThroughUrl':{'customClickThroughUrl':url_list[0]},'richMediaExitOverrides':[{'exitId':exitid_list[0],'enabled':True,'clickThroughUrl':{'customClickThroughUrl':url_list[0]}}]}]},'active':True}).execute()
        elif len(creative_list)==2: #If there are two creatives on rotation
            dcm_create_ad = DFA.ads().insert(profileId=profile_id,body={'campaignId':campaign_id,'advertiserId':advertiser_id,'name':ad_name,'startTime':dcm_placement_startdate+'T05:00:00Z','endTime':dcm_placement_enddate+'T00:00:00Z','type':'AD_SERVING_STANDARD_AD','deliverySchedule':{'priority':'AD_PRIORITY_13','impressionRatio':1},'placementAssignments':[{'active':True,'placementId':dcm_placement_id}],'creativeRotation':{'creativeAssignments':[{'active':True,'creativeId':creative_list[0],'clickThroughUrl':{'customClickThroughUrl':url_list[0]},'richMediaExitOverrides':[{'exitId':exitid_list[0],'enabled':True,'clickThroughUrl':{'customClickThroughUrl':url_list[0]}}]},{'active':True,'creativeId':creative_list[1],'clickThroughUrl':{'customClickThroughUrl':url_list[1]},'richMediaExitOverrides':[{'exitId':exitid_list[1],'enabled':True,'clickThroughUrl':{'customClickThroughUrl':url_list[1]}}]}]},'active':True}).execute()
        elif len(creative_list)==3: #If there are three creatives on rotation
            dcm_create_ad = DFA.ads().insert(profileId=profile_id,body={'campaignId':campaign_id,'advertiserId':advertiser_id,'name':ad_name,'startTime':dcm_placement_startdate+'T05:00:00Z','endTime':dcm_placement_enddate+'T00:00:00Z','type':'AD_SERVING_STANDARD_AD','deliverySchedule':{'priority':'AD_PRIORITY_13','impressionRatio':1},'placementAssignments':[{'active':True,'placementId':dcm_placement_id}],'creativeRotation':{'creativeAssignments':[{'active':True,'creativeId':creative_list[0],'clickThroughUrl':{'customClickThroughUrl':url_list[0]},'richMediaExitOverrides':[{'exitId':exitid_list[0],'enabled':True,'clickThroughUrl':{'customClickThroughUrl':url_list[0]}}]},{'active':True,'creativeId':creative_list[1],'clickThroughUrl':{'customClickThroughUrl':url_list[1]},'richMediaExitOverrides':[{'exitId':exitid_list[1],'enabled':True,'clickThroughUrl':{'customClickThroughUrl':url_list[1]}}]},{'active':True,'creativeId':creative_list[2],'clickThroughUrl':{'customClickThroughUrl':url_list[2]},'richMediaExitOverrides':[{'exitId':exitid_list[2],'enabled':True,'clickThroughUrl':{'customClickThroughUrl':url_list[2]}}]}]},'active':True}).execute()

        print('Finished processing row ' + str(j))
        print('\n')
        sheet['A'+str(j)].font = Font(bold=True)
        workbook.save(config.feed_path)
    except Exception as e: #If there is an error, skip the row, display the error and go to the next line
        print('###Skipped row ' + str(j))
        print(e)
        skip_count += 1
        continue

print('Reached the end')
print('Skipped a total of ' + str(skip_count) + ' rows')
print("--- Time taken - %s seconds ---" % (time.time() - start_time))
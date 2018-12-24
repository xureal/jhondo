def gvalidate(filename, filedir, creds):
	import openpyxl
	import gspread
	from oauth2client.service_account import ServiceAccountCredentials
	import re


	scope = ['https://spreadsheets.google.com/feeds']
	credentials = ServiceAccountCredentials.from_json_keyfile_name(creds, scope)
	gc = gspread.authorize(credentials)


	#Excel file
	#wb = openpyxl.load_workbook(r'D:\Users\andptasi\Desktop\Generic Template.xlsm')
	wb = openpyxl.load_workbook(filename)
	#wbData = openpyxl.load_workbook(r'D:\Users\andptasi\Desktop\Generic Template.xlsm', data_only=True)
	wbData = openpyxl.load_workbook(filename, data_only=True)
	sheet = wb['Media Plan']
	sheetData = wbData['Media Plan']
	campaignSheetData = wbData['Campaign Spreadsheet']
	count = sheet.max_row
	cscount = campaignSheetData.max_row



	#Google Sheets
	sitelistFile = gc.open_by_key('1AuVnBS6M0W9LXmq5m2sJ7-po04zyLQBV3MjbOAJ_rlk')
	sitelistSheet = sitelistFile.get_worksheet(6)

	#Collect list of countries from the google sheet and put it into a set
	setOfCountries = set(sitelistSheet.col_values(1))	
	
	#Collect list of campaign codes from the google sheet and put it into a set
	setOfCampaignCode = set(sitelistSheet.col_values(4))
	
	#Collect list of merchant codes from the google sheet and put it into a set
	setOfMerchantCode = set(sitelistSheet.col_values(7))

	#Collect list of publishers from the google sheet and put it into a set
	setOfPublishers = set(sitelistSheet.col_values(10))
	
	#Collect list of targets from the google sheet and put it into a set
	setOfTarget = set(sitelistSheet.col_values(13))
	
	#Collect list of devices from the google sheet and put it into a set
	setOfDevice = set(sitelistSheet.col_values(16))

	#Collect list of placement type from the google sheet and put it into a set
	setOfPlacementType = set(sitelistSheet.col_values(19))
	
	#Collect list of placement rate from the google sheet and put it into a set
	setOfPlacementRate = set(sitelistSheet.col_values(23))
	
	#Collect list of placement size from the google sheet and put it into a set
	setOfPlacementSize = set(sitelistSheet.col_values(25))
	
	#Collect list of ad serving method size from the google sheet and put it into a set
	setOfAdServingMethod = set(sitelistSheet.col_values(28))
	
	#Collect list of ad type size from the google sheet and put it into a set
	setOfAdType = set(sitelistSheet.col_values(31))

	#Collect list of creative asset from the google sheet and put it into a set
	setOfCreativeAsset = set(sitelistSheet.col_values(34))
	
	#DO NOT USE THE TAG TYPE#
	#Collect list of tag type from the google sheet and put it into a set
	#setOfTagType = set(sitelistSheet.col_values(37))
	
	#Collect list of vendor tag from the google sheet and put it into a set
	setOfVendorTag = set(sitelistSheet.col_values(40))
	
	#Collect list of firing rule from the google sheet and put it into a set
	setOfFiringRule = set(sitelistSheet.col_values(43))
	
	#Collect list of offer type from the google sheet and put it into a set
	setOfOfferType = set(sitelistSheet.col_values(46))

	#Collect list of date from the google sheet and put it into a set
	setOfDate = set(sitelistSheet.col_values(47))
	
	#Collect list of campaign objective from the google sheet and put it into a set
	setOfCampaignObjective = set(sitelistSheet.col_values(48))
	
	#Collect list of media channel from the google sheet and put it into a set
	setOfMediaChannel = set(sitelistSheet.col_values(51))
	
	#Collect list of inventory type from the google sheet and put it into a set
	setOfInventoryType = set(sitelistSheet.col_values(54))
	
	#Collect list of verification partner size from the google sheet and put it into a set
	setOfVerificationPartner = set(sitelistSheet.col_values(57))

	#Collect list of creative type from the google sheet and put it into a set
	setOfCreativeType = set(sitelistSheet.col_values(60))
	
	#Collect list of tag type from the google sheet and put it into a set
	setOfTagType = set(sitelistSheet.col_values(61))

	#Collect list of format from the google sheet and put it into a set
	setOfFormat = set(sitelistSheet.col_values(62))
	
	#Collect list of language from the google sheet and put it into a set
	setOfLanguage = set(sitelistSheet.col_values(63))

	#Regex
	sizePattern = re.compile("^[0-9]{1,4}(x)[0-9]{1,4}$")
	urlPattern = re.compile("^(https?|ftp):\/\/(-\.)?([^\s\/?\.]+\.?)+(\/[^\s]*)?$")
	datePattern = re.compile("^(0[1-9]|1[0-2])\/(0[1-9]|1\d|2\d|3[01])\/[2][0][1-2][0-9]$")
	placementPattern = re.compile("^[a-zA-Z0-9!@#$()\\-.+ ]*$")
	creativePattern = re.compile("^[a-zA-z0-9]{2,}(_)[0-9]{1,4}(x)[0-9]{1,4}(_)[a-zA-z]{2,3}$")

	logFile = open(filedir+"\\TS_results.txt", 'w')
	logString = ''
	csPlcList = []
	errorFlag = 0
	warningFlag = 0

	cmpmarket = str(sheet.cell(row=3, column=2).value)
	merchantcode = str(sheet.cell(row=5, column=2).value)
	cmpdate = str(sheet.cell(row=6, column=2).value)
	cmpobjective = str(sheet.cell(row=8, column=2).value)

	if cmpmarket!=None and cmpmarket!='':
		if cmpmarket not in setOfCountries:
			print(" The campaign market - " + cmpmarket + "  is not valid, please revise it!")
			logFile.write(" The campaign market - " + cmpmarket + "  is not valid, please revise it!\n")
			logString+=(" The campaign market - " + cmpmarket + "  is not valid, please revise it!\n")
			errorFlag = 1
			
	if merchantcode!=None and merchantcode!='':
		if merchantcode not in setOfMerchantCode:
			print(" The merchant code - " + merchantcode + "  is not valid, please revise it!")
			logFile.write(" The merchant code - " + merchantcode + "  is not valid, please revise it!\n")
			logString+=(" The merchant code - " + merchantcode + "  is not valid, please revise it!\n")
			errorFlag = 1			
			
	if cmpdate!=None and cmpdate!='':
		if cmpdate not in setOfDate:
			print(" The campaign date - " + cmpdate + "  is not valid, please revise it!")
			logFile.write(" The campaign date - " + cmpdate + "  is not valid, please revise it!\n")
			logString+=(" The campaign date - " + cmpdate + "  is not valid, please revise it!\n")
			errorFlag = 1						

	if cmpobjective!=None and cmpobjective!='':
		if cmpobjective not in setOfCampaignObjective:
			print(" The campaign objective - " + cmpobjective + "  is not valid, please revise it!")
			logFile.write(" The campaign objective - " + cmpobjective + "  is not valid, please revise it!\n")
			logString+=(" The campaign objective - " + cmpobjective + "  is not valid, please revise it!\n")
			errorFlag = 1							
			

	#Cycle through the campaign details
	for i in range (3,10,1):
		cmpDetails = str(sheet.cell(row=i, column=2).value)
		if cmpDetails!=None and cmpDetails!='':
			if cmpDetails.strip() != cmpDetails:
				print("The " + str(sheet.cell(row=i,column=1).value) + " field has leading or trailing spaces!")
				logFile.write("The " + str(sheet.cell(row=i,column=1).value) + " field has leading or trailing spaces!\n")
				logString+=("The " + str(sheet.cell(row=i,column=1).value) + " field has leading or trailing spaces!\n")
				errorFlag = 1
		


	#Cycle through all of the rows on the excel file
	for i in range (12,count,1):

		#Cycle through all of the columns on the excel file
		for j in range (1,31,1):
			#Check for leading and trailing spaces
			cellValue = str(sheet.cell(row=i,column=j).value)
			if cellValue!=None and cellValue!='':
				if cellValue.strip() != cellValue:
					print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " has leading or trailing spaces!")
					logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " has leading or trailing spaces!\n")
					logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " has leading or trailing spaces!\n")
					errorFlag = 1
			#Check if publisher is there but other columns are blank
			publisher = sheet.cell(row=i,column=1).value		
			if publisher!=None and publisher!='':
				if sheet.cell(row=i,column=j).value==None or sheet.cell(row=i,column=j).value=='':
					if (j==21 or j==22 or j==23 or j==24) and ((sheet.cell(row=i,column=19).value==None or sheet.cell(row=i,column=19).value=='') and (sheet.cell(row=i,column=20).value==None or sheet.cell(row=i,column=20).value=='')):
						print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!")
						logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
						logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
						errorFlag = 1
					elif (j!=21 and j!=22 and j!=23 and j!=24):
						print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!")
						logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
						logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
						errorFlag = 1
						
		# Check publisher column
		# Check if it's present in the publisher list on google sheet
		if publisher!=None and publisher!='':
			if publisher not in setOfPublishers:
				print("In row " + str(i) + " the publisher - " + publisher + " is not in the list of defined publishers!")
				logFile.write("In row " + str(i) + " the publisher - " + publisher + " is not in the list of defined publishers!\n")
				logString+=("In row " + str(i) + " the publisher - " + publisher + " is not in the list of defined publishers!\n")
				errorFlag = 1

		#Check size column
		#Match it to regex above of 1-4 numbers, lowercase x, 1-4 numbers
		size = sheet.cell(row=i,column=14).value
		if size!=None and size!='':
			if size not in setOfPlacementSize:
				print("In row " + str(i) + " the size - " + size + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the size - " + size + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the size - " + size + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check language column
		#Check if it's a 3 letter code present on the google sheet
		language = sheet.cell(row=i,column=12).value
		if language!=None and language!='':
			if language not in setOfLanguage:
				print("In row " + str(i) + " the language - " + language + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the language - " + language + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the language - " + language + " is not valid, please revise it!\n")
				errorFlag = 1

		#Check country column
		#Check if it's a 2 letter ISO country code that's on the google sheet
		country = sheet.cell(row=i,column=11).value
		if country!=None and country!='':
			if country not in setOfCountries:
				print("In row " + str(i) + " the country - " + country + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the country - " + country + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the country - " + country + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check creative type column
		#Check if it's present in the creative type list on google sheet
		creativeType = sheet.cell(row=i,column=15).value
		if creativeType!=None and creativeType!='':
			if creativeType not in setOfCreativeType:
				print("In row " + str(i) + " the creative type - " + creativeType + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the creative type - " + creativeType + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the creative type - " + creativeType + " is not valid, please revise it!\n")
				errorFlag = 1

		#Check device column
		#Check if it's present in the device list on google sheet
		device = sheet.cell(row=i,column=16).value
		if device!=None and device!='':
			if device not in setOfDevice:
				print("In row " + str(i) + " the device - " + device + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the device - " + device + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the device - " + device + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check cost model column
		#Check if it's present in the cost model list on google sheet
		costModel = sheet.cell(row=i,column=27).value
		if costModel!=None and costModel!='':
			if costModel not in setOfPlacementRate:
				print("In row " + str(i) + " the cost model - " + costModel + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the cost model - " + costModel + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the cost model - " + costModel + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check tag type column
		#Check if it's present in the tag type list on google sheet
		tagType = sheet.cell(row=i,column=13).value
		if tagType!=None and tagType!='':
			if tagType not in setOfTagType:
				print("In row " + str(i) + " the tag type - " + tagType + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the tag type - " + tagType + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the tag type - " + tagType + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check ad server column
		#Check if it's present in the ad server list on google sheet
		adServer = sheet.cell(row=i,column=9).value
		if adServer!=None and adServer!='':
			if adServer not in setOfAdServingMethod:
				print("In row " + str(i) + " the ad server - " + adServer + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the ad server - " + adServer + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the ad server - " + adServer + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check objective column
		#Check if it's present in the objective list on google sheet
		# objective = sheet.cell(row=i,column=3).value
		# if objective!=None and objective!='':
			# if objective not in setOfObjective:
				# print "In row " + str(i) + " the objective - " + objective + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the objective - " + objective + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the objective - " + objective + " is not valid, please revise it!\n")
				# errorFlag = 1
				
		#Check format column
		#Check if it's present in the format list on google sheet
		format = sheet.cell(row=i,column=10).value
		if format!=None and format!='':
			if format not in setOfFormat:
				print("In row " + str(i) + " the format - " + format + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the format - " + format + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the format - " + format + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check channel column
		#Check if it's present in the channel list on google sheet
		channel = sheet.cell(row=i,column=3).value
		if channel!=None and channel!='':
			if channel not in setOfMediaChannel:
				print("In row " + str(i) + " the channel - " + channel + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the channel - " + channel + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the channel - " + channel + " is not valid, please revise it!\n")
				errorFlag = 1

		#Check placement type column
		#Check if it's present in the placement type list on google sheet
		placementType = sheet.cell(row=i,column=4).value
		if placementType!=None and placementType!='':
			if placementType not in setOfPlacementType:
				print("In row " + str(i) + " the placement type - " + placementType + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the placement type - " + placementType + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the placement type - " + placementType + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check target column
		#Check if it's present in the target list on google sheet
		target = sheet.cell(row=i,column=5).value
		if target!=None and target!='':
			if target not in setOfTarget:
				print("In row " + str(i) + " the target - " + target + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the target - " + target + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the target - " + target + " is not valid, please revise it!\n")
				errorFlag = 1				
				
		#Check ad type column
		#Check if it's present in the ad type list on google sheet
		adType = sheet.cell(row=i,column=7).value
		if adType!=None and adType!='':
			if adType not in setOfAdType:
				print("In row " + str(i) + " the ad type - " + adType + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the ad type - " + adType + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the ad type - " + adType + " is not valid, please revise it!\n")
				errorFlag = 1							
				
		#Check inventory type column
		#Check if it's present in the inventory type list on google sheet
		inventoryType = sheet.cell(row=i,column=17).value
		if inventoryType!=None and inventoryType!='':
			if inventoryType not in setOfInventoryType:
				print("In row " + str(i) + " the inventory type - " + inventoryType + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the inventory type - " + inventoryType + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the inventory type - " + inventoryType + " is not valid, please revise it!\n")
				errorFlag = 1

		#Check vvf column
		#Check if it's present in the vvf list on google sheet
		vvf = sheet.cell(row=i,column=18).value
		if vvf!=None and vvf!='':
			if vvf not in setOfVerificationPartner:
				print("In row " + str(i) + " the vvf - " + vvf + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " vvf - " + vvf + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " vvf - " + vvf + " is not valid, please revise it!\n")
				errorFlag = 1			

		# #Check cost structure column
		# #Check if it's present in the cost structure list on google sheet
		# costStructure = sheet.cell(row=i,column=17).value
		# if costStructure!=None and costStructure!='':
			# if costStructure not in setOfCostStructure:
				# print "In row " + str(i) + " the cost structure - " + costStructure + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the cost structure - " + costStructure + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the cost structure - " + costStructure + " is not valid, please revise it!\n")
				# errorFlag = 1
				
		#Check URL column
		#Match it to the regex above for URL matching
		url1 = sheet.cell(row=i,column=20).value
		url2 = sheet.cell(row=i,column=22).value
		url3 = sheet.cell(row=i,column=24).value
		if url1!=None and url1!='':
			if not urlPattern.match(url1):
				print("In row " + str(i) + " url1 - " + url1 + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " url1 - " + url1 + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " url1 - " + url1 + " is not valid, please revise it!\n")
				errorFlag = 1
		if url2!=None and url2!='':
			if not urlPattern.match(url2):
				print("In row " + str(i) + " url2 - " + url2 + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " url2 - " + url2 + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " url2 - " + url2 + " is not valid, please revise it!\n")
				errorFlag = 1
		if url3!=None and url3!='':
			if not urlPattern.match(url3):
				print("In row " + str(i) + " url3 - " + url3 + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " url3 - " + url3 + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " url3 - " + url3 + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check date column
		#Match it to regex above of mm/dd/yyyy
		sDate = sheet.cell(row=i,column=25).value
		eDate = sheet.cell(row=i,column=26).value


		# if sDate!=None and sDate!='':
			# startDate = sDate.strftime('%m/%d/%Y')
			# if not datePattern.match(startDate):
				# print "In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!\n")
				# errorFlag = 1
		# if eDate!=None and eDate!='':
			# endDate = eDate.strftime('%m/%d/%Y')
			# if not datePattern.match(endDate):
				# print "In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!\n")
				# errorFlag = 1
				
			
		if sDate!=None and sDate!='':
			try:
				startDate = sDate.strftime('%m/%d/%Y')
			except:
				startDate = ''
			if not datePattern.match(startDate):
				print("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!\n")
				errorFlag = 1
		if eDate!=None and eDate!='':
			try:
				endDate = eDate.strftime('%m/%d/%Y')
			except:
				endDate = ''
			if not datePattern.match(endDate):
				print("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!\n")
				errorFlag = 1				
				

		#Check total USD column
		#Flag a warning if cost is above threshold
		totalCost = sheetData.cell(row=i,column=24).value
		# if totalCost !=None and totalCost!='':
			# if totalCost > 2000:
				# print "In row " + str(i) + " the total cost is - " + str(totalCost) + ". Are you sure this is correct?"
				# logFile.write("In row " + str(i) + " the total cost is - " + str(totalCost) + ". Are you sure this is correct?\n")
				# logString+=("In row " + str(i) + " the total cost is - " + str(totalCost) + ". Are you sure this is correct?\n")
				# warningFlag = 1
				
		#Check the placement name column
		#Check it to match the special characer regex above
		plcName = sheet.cell(row=i,column=2).value
		if plcName!=None and plcName!='':
			if not placementPattern.match(str(plcName)):
				print("In row " + str(i) + " the placement name - " + plcName + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the placement name - " + plcName + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the placement name - " + plcName + " is not valid, please revise it!\n")
				errorFlag = 1

		#Check the target description column
		#Check it to match the special characer regex above
		targetdesc = sheet.cell(row=i,column=6).value
		if targetdesc!=None and targetdesc!='':
			if not placementPattern.match(str(targetdesc)):
				print("In row " + str(i) + " the target description - " + targetdesc + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the target description - " + targetdesc + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the target description - " + targetdesc + " is not valid, please revise it!\n")
				errorFlag = 1				

				
		#Check the ad description name column
		#Check it to match the special characer regex above
		addesc = sheet.cell(row=i,column=2).value
		if addesc!=None and addesc!='':
			if not placementPattern.match(str(addesc)):
				print("In row " + str(i) + " the ad description - " + addesc + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the ad description - " + addesc + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the ad description - " + addesc + " is not valid, please revise it!\n")
				errorFlag = 1
				
				
				

		#Check the creative name column
		#Check if it matches the naming convention (client_creative_size_language)
		# crtName1 = sheet.cell(row=i,column=12).value
		# crtName2 = sheet.cell(row=i,column=14).value
		# crtName3 = sheet.cell(row=i,column=16).value
		# if crtName1!=None and crtName1!='' and crtName1.lower()!='tracking':
			# if not creativePattern.match(crtName1):
				# print "In row " + str(i) + " the creative name - " + crtName1 + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the creative name - " + crtName1 + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the creative name - " + crtName1 + " is not valid, please revise it!\n")
				# errorFlag = 1
		# if crtName2!=None and crtName2!='' and crtName2.lower()!='tracking':
			# if not creativePattern.match(crtName2):
				# print "In row " + str(i) + " the creative name - " + crtName2 + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the creative name - " + crtName2 + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the creative name - " + crtName2 + " is not valid, please revise it!\n")
				# errorFlag = 1
		# if crtName3!=None and crtName3!='' and crtName3.lower()!='tracking':
			# if not creativePattern.match(crtName3):
				# print "In row " + str(i) + " the creative name - " + crtName3 + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the creative name - " + crtName3 + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the creative name - " + crtName3 + " is not valid, please revise it!\n")
				# errorFlag = 1
			
	#Check the placement name column on the Campaign Spreadsheet
	#Check to see if there are duplicate placements
	#Go through each line on the CS sheet
	for k in range (2,cscount,1):
		csPlcName = campaignSheetData.cell(row=k,column=4).value
		if csPlcName!=None and csPlcName!='':
			csPlcList.append(csPlcName)	
	if len(csPlcList) != len(set(csPlcList)):
		print("There are duplicate placements - please check the Campaign Spreadsheet tab for the highlighted rows")
		logFile.write("There are duplicate placements - please check the Campaign Spreadsheet tab for the highlighted rows\n")
		logString+=("There are duplicate placements - please check the Campaign Spreadsheet tab for the highlighted rows\n")
		errorFlag = 1
			

	logFile.close()
	return errorFlag, warningFlag, logString
	

def gvalidate(filename, filedir, creds):
	import openpyxl
	import gspread
	from oauth2client.service_account import ServiceAccountCredentials
	import re


	scope = ['https://spreadsheets.google.com/feeds']
	credentials = ServiceAccountCredentials.from_json_keyfile_name(creds, scope)
	gc = gspread.authorize(credentials)


	#Excel file
	#wb = openpyxl.load_workbook(r'D:\Users\andptasi\Desktop\Generic Template.xlsm')
	wb = openpyxl.load_workbook(filename)
	#wbData = openpyxl.load_workbook(r'D:\Users\andptasi\Desktop\Generic Template.xlsm', data_only=True)
	wbData = openpyxl.load_workbook(filename, data_only=True)
	sheet = wb['Media Plan']
	sheetData = wbData['Media Plan']
	campaignSheetData = wbData['Campaign Spreadsheet']
	count = sheet.max_row
	cscount = campaignSheetData.max_row



	#Google Sheets
	sitelistFile = gc.open_by_key('1AuVnBS6M0W9LXmq5m2sJ7-po04zyLQBV3MjbOAJ_rlk')
	sitelistSheet = sitelistFile.get_worksheet(3)

	#Collect list of publishers from the google sheet and put it into a set
	setOfPublishers = set(sitelistSheet.col_values(1))

	#Collect list of languages from the google sheet and put it into a set
	setOfLanguages = set(sitelistSheet.col_values(7))

	#Collect list of countries from the google sheet and put it into a set
	setOfCountries = set(sitelistSheet.col_values(6))

	#Collect list of creative types from the google sheet and put it into a set
	setOfCreativeTypes = set(sitelistSheet.col_values(9))

	#Collect list of Campaign Category from the google sheet and put it into a set
	setOfCampaignCategory = set(sitelistSheet.col_values(13))

	#Collect list of cost model from the google sheet and put it into a set
	setOfCostModel = set(sitelistSheet.col_values(11))

	#Collect list of campaign type from the google sheet and put it into a set
	setOfCampaignType = set(sitelistSheet.col_values(12))
	
	#Collect list of tag type from the google sheet and put it into a set
	setOfTagType = set(sitelistSheet.col_values(8))
	
	#Collect list of gender from the google sheet and put it into a set
	setOfGender = set(sitelistSheet.col_values(10))
	
	#Collect list of targeting from the google sheet and put it into a set
	setOfTargeting= set(sitelistSheet.col_values(2))
	
	#Collect list of format from the google sheet and put it into a set
	setOfFormat = set(sitelistSheet.col_values(3))
	
	#Collect list of channel from the google sheet and put it into a set
	setOfChannel = set(sitelistSheet.col_values(4))
	
	#Collect list of campaign strategy from the google sheet and put it into a set
	setOfCampaignStrategy = set(sitelistSheet.col_values(14))	

	#Collect list of campaign date from the google sheet and put it into a set
	setOfCampaignDate = set(sitelistSheet.col_values(15))	

	#Collect list of platform from the google sheet and put it into a set
	setOfPlatforms = set(sitelistSheet.col_values(17))	

	#Collect list of customer type from the google sheet and put it into a set
	setOfCustomerTypes = set(sitelistSheet.col_values(18))	

	#Regex
	sizePattern = re.compile("^[0-9]{1,4}(x)[0-9]{1,4}$")
	urlPattern = re.compile("^(https?|ftp):\/\/(-\.)?([^\s\/?\.#-]+\.?)+(\/[^\s]*)?$")
	datePattern = re.compile("^(0[1-9]|1[0-2])\/(0[1-9]|1\d|2\d|3[01])\/[2][0][1-2][0-9]$")
	placementPattern = re.compile("^[a-zA-Z0-9!@#$()\\-.+_ ]*$")
	creativePattern = re.compile("^[a-zA-z0-9]{2,}(_)[0-9]{1,4}(x)[0-9]{1,4}(_)[a-zA-z]{2,3}$")

	logFile = open(filedir+"\\TS_results.txt", 'w')
	logString = ''
	csPlcList = []
	errorFlag = 0
	warningFlag = 0

	#Cycle through the campaign details
	for i in range (3,10,1):
		cmpDetails = str(sheet.cell(row=i, column=2).value)
		if cmpDetails!=None and cmpDetails!='':
			if cmpDetails.strip() != cmpDetails:
				print("The " + str(sheet.cell(row=i,column=1).value) + " field has leading or trailing spaces!")
				logFile.write("The " + str(sheet.cell(row=i,column=1).value) + " field has leading or trailing spaces!\n")
				logString+=("The " + str(sheet.cell(row=i,column=1).value) + " field has leading or trailing spaces!\n")
				errorFlag = 1
				#Cycle through all of the rows on the excel file
	for i in range (12,count,1):
		#Check size column to see if the row is a package
		size = sheet.cell(row=i,column=9).value
		if size == 'Package' or size == 'package':
			packagePublisher = sheet.cell(row=i,column=1).value
			nextRowPublisher = sheet.cell(row=i+1,column=1).value
			if packagePublisher != nextRowPublisher:
				print("In row " + str(i) + " the package publisher doesn't match the placement publisher")
				logFile.write("In row " + str(i) + " the package publisher doesn't match the placement publisher\n")
				logString+=("In row " + str(i) + " the package publisher doesn't match the placement publisher\n")
				errorFlag = 1
			packageDescription = sheet.cell(row=i,column=2).value
			if not placementPattern.match(str(packageDescription)) or packageDescription == '' or packageDescription == None:
				print("In row " + str(i) + " the package description is not valid")
				logFile.write("In row " + str(i) + " the package description is not valid\n")
				logString+=("In row " + str(i) + " the package description is not valid\n")
				errorFlag = 1
			packageMarket = sheet.cell(row=i,column=6).value
			if packageMarket not in setOfCountries or packageMarket == '' or packageMarket == None:
				print("In row " + str(i) + " the package market is not valid")
				logFile.write("In row " + str(i) + " the package market is not valid\n")
				logString+=("In row " + str(i) + " the package market is not valid\n")
				errorFlag = 1
			package_sDate = sheet.cell(row=i,column=17).value
			package_eDate = sheet.cell(row=i,column=18).value
			if package_sDate!=None and package_sDate!='':
				pstartDate = package_sDate.strftime('%m/%d/%Y')
				if not datePattern.match(pstartDate):
					print("In row " + str(i) + " the package start date is not valid")
					logFile.write("In row " + str(i) + " the package start date is not valid\n")
					logString+=("In row " + str(i) + " the package start date is not valid\n")
					errorFlag = 1
			if package_eDate!=None and package_eDate!='':
				pendDate = package_eDate.strftime('%m/%d/%Y')
				if not datePattern.match(pendDate):
					print("In row " + str(i) + " the package end date is not valid")
					logFile.write("In row " + str(i) + " the package end date is not valid\n")
					logString+=("In row " + str(i) + " the package end date is not valid\n")
					errorFlag = 1
			packageCostModel = sheet.cell(row=i,column=19).value
			if packageCostModel not in setOfCostModel or packageCostModel == '' or packageCostModel == None:
				print("In row " + str(i) + " the package cost model is not valid")
				logFile.write("In row " + str(i) + " the package cost model is not valid\n")
				logString+=("In row " + str(i) + " the package cost model is not valid\n")
				errorFlag = 1
			packageUnits = sheet.cell(row=i,column=21).value
			packageRate = sheet.cell(row=i,column=22).value
			if packageUnits == '' or packageUnits == None:
				print("In row " + str(i) + " the package units are blank")
				logFile.write("In row " + str(i) + " the package units are blank\n")
				logString+=("In row " + str(i) + " the package units are blank\n")
				errorFlag = 1
			if packageRate == '' or packageRate == None:
				print("In row " + str(i) + " the package rate is blank")
				logFile.write("In row " + str(i) + " the package rate is blank\n")
				logString+=("In row " + str(i) + " the package rate is blank\n")
				errorFlag = 1

		#Cycle through all of the columns on the excel file
		for j in range (1,23,1):
			if size == 'Package' or size == 'package':
				if not(sheet.cell(row=i,column=j).value==None or sheet.cell(row=i,column=j).value=='') and (j!=1 and j!=2 and j!=6 and j!=9 and j!=17 and j!=18 and j!=19 and j!=21 and j!=22 and j!=23 and j!=24):
					print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " must be blank if the line is a package")
					logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " must be blank if the line is a package\n")
					logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " must be blank if the line is a package\n")
					errorFlag = 1

			#Check for leading and trailing spaces
			cellValue = str(sheet.cell(row=i,column=j).value)
			if cellValue!=None and cellValue!='':
				if cellValue.strip() != cellValue:
					print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " has leading or trailing spaces!")
					logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " has leading or trailing spaces!\n")
					logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " has leading or trailing spaces!\n")
					errorFlag = 1

			#Check if publisher is there but other columns are blank
			publisher = sheet.cell(row=i,column=1).value
			size = sheet.cell(row=i,column=9).value			
			if (publisher!=None and publisher!='') and not(size=='Package' or size=='package'):
				if sheet.cell(row=i,column=j).value==None or sheet.cell(row=i,column=j).value=='':
					if (j==13 or j==14 or j==15 or j==16) and ((sheet.cell(row=i,column=11).value==None or sheet.cell(row=i,column=11).value=='') and (sheet.cell(row=i,column=13).value==None or sheet.cell(row=i,column=13).value=='')):
						print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!")
						logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
						logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
						errorFlag = 1
					elif (j!=13 and j!=14 and j!=15 and j!=16 and j!=21 and j!=22):
						print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!")
						logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
						logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
						errorFlag = 1
						
		# Check publisher column
		# Check if it's present in the publisher list on google sheet
		publisher = sheet.cell(row=i,column=1).value
		if publisher!=None and publisher!='':
			if publisher not in setOfPublishers:
				print("In row " + str(i) + " the publisher - " + publisher + " is not in the list of defined publishers!")
				logFile.write("In row " + str(i) + " the publisher - " + publisher + " is not in the list of defined publishers!\n")
				logString+=("In row " + str(i) + " the publisher - " + publisher + " is not in the list of defined publishers!\n")
				errorFlag = 1

				
		#Check language column
		#Check if it's a 3 letter code present on the google sheet
		language = sheet.cell(row=i,column=7).value
		if language!=None and language!='':
			if language not in setOfLanguages:
				print("In row " + str(i) + " the language - " + language + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the language - " + language + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the language - " + language + " is not valid, please revise it!\n")
				errorFlag = 1
		#Check country column
		#Check if it's a 2 letter ISO country code that's on the google sheet
		country = sheet.cell(row=i,column=6).value
		if country!=None and country!='':
			if country not in setOfCountries:
				print("In row " + str(i) + " the country - " + country + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the country - " + country + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the country - " + country + " is not valid, please revise it!\n")
				errorFlag = 1

		#Check creative type column
		#Check if it's present in the creative type list on google sheet
		creativeType = sheet.cell(row=i,column=10).value
		if creativeType!=None and creativeType!='':
			if creativeType not in setOfCreativeTypes:
				print("In row " + str(i) + " the creative type - " + creativeType + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the creative type - " + creativeType + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the creative type - " + creativeType + " is not valid, please revise it!\n")
				errorFlag = 1

		#Check platform column
		#Check if it's present in the platform list on google sheet
		# platform = sheet.cell(row=i,column=12).value
		# if platform!=None and platform!='':
			# if platform not in setOfPlatforms:
				# print "In row " + str(i) + " the platform - " + platform + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the platform - " + platform + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the platform - " + platform + " is not valid, please revise it!\n")
				# errorFlag = 1
				
		#Check customer type column
		#Check if it's present in the customer type list on google sheet
		# customerType = sheet.cell(row=i,column=13).value
		# if customerType!=None and customerType!='':
			# if customerType not in setOfCustomerTypes:
				# print "In row " + str(i) + " the customer type - " + customerType + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the customer type - " + customerType + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the customer type - " + customerType + " is not valid, please revise it!\n")
				# errorFlag = 1
				
		#Check device column
		#Check if it's present in the device list on google sheet
		# gender = sheet.cell(row=i,column=11).value
		# if gender!=None and gender!='':
			# if gender not in setOfGender:
				# print "In row " + str(i) + " the gender - " + gender + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the gender - " + gender + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the gender - " + gender + " is not valid, please revise it!\n")
				# errorFlag = 1
				
		#Check cost model column
		#Check if it's present in the cost model list on google sheet
		costModel = sheet.cell(row=i,column=19).value
		if costModel!=None and costModel!='':
			if costModel not in setOfCostModel:
				print("In row " + str(i) + " the cost model - " + costModel + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the cost model - " + costModel + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the cost model - " + costModel + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check tag type column
		#Check if it's present in the tag type list on google sheet
		tagType = sheet.cell(row=i,column=8).value
		if tagType!=None and tagType!='':
			if tagType not in setOfTagType:
				print("In row " + str(i) + " the tag type - " + tagType + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the tag type - " + tagType + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the tag type - " + tagType + " is not valid, please revise it!\n")
				errorFlag = 1
				

		#Check objective column
		#Check if it's present in the objective list on google sheet
		targeting = sheet.cell(row=i,column=3).value
		if targeting!=None and targeting!='':
			if targeting not in setOfTargeting:
				print("In row " + str(i) + " the targeting - " + targeting + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the targeting - " + targeting + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the targeting - " + targeting + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check format column
		#Check if it's present in the format list on google sheet
		# format = sheet.cell(row=i,column=4).value
		# if format!=None and format!='':
			# if format not in setOfFormat:
				# print "In row " + str(i) + " the format - " + format + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the format - " + format + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the format - " + format + " is not valid, please revise it!\n")
				# errorFlag = 1
				
		#Check channel column
		#Check if it's present in the channel list on google sheet
		channel = sheet.cell(row=i,column=5).value
		if channel!=None and channel!='':
			if channel not in setOfChannel:
				print("In row " + str(i) + " the channel - " + channel + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the channel - " + channel + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the channel - " + channel + " is not valid, please revise it!\n")
				errorFlag = 1

		#Check URL column
		#Match it to the regex above for URL matching
		url1 = sheet.cell(row=i,column=12).value
		url2 = sheet.cell(row=i,column=14).value
		url3 = sheet.cell(row=i,column=16).value
		if url1!=None and url1!='':
			if not urlPattern.match(url1):
				print("In row " + str(i) + " url1 - " + url1 + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " url1 - " + url1 + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " url1 - " + url1 + " is not valid, please revise it!\n")
				errorFlag = 1
		if url2!=None and url2!='':
			if not urlPattern.match(url2):
				print("In row " + str(i) + " url2 - " + url2 + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " url2 - " + url2 + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " url2 - " + url2 + " is not valid, please revise it!\n")
				errorFlag = 1
		if url3!=None and url3!='':
			if not urlPattern.match(url3):
				print("In row " + str(i) + " url3 - " + url3 + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " url3 - " + url3 + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " url3 - " + url3 + " is not valid, please revise it!\n")
				errorFlag = 1
				
		#Check date column
		#Match it to regex above of mm/dd/yyyy
		sDate = sheet.cell(row=i,column=17).value
		eDate = sheet.cell(row=i,column=18).value


		if sDate!=None and sDate!='':
			try:
				startDate = sDate.strftime('%m/%d/%Y')
			except:
				startDate = ''
			if not datePattern.match(startDate):
				print("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!\n")
				errorFlag = 1
		if eDate!=None and eDate!='':
			try:
				endDate = eDate.strftime('%m/%d/%Y')
			except:
				endDate = ''
			if not datePattern.match(endDate):
				print("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!\n")
				errorFlag = 1


		#Check total USD column
		#Flag a warning if cost is above threshold
		totalCost = sheetData.cell(row=i,column=22).value
		if totalCost !=None and totalCost!='':
			if totalCost > 2000:
				print("In row " + str(i) + " the total cost is - " + str(totalCost) + ". Are you sure this is correct?")
				logFile.write("In row " + str(i) + " the total cost is - " + str(totalCost) + ". Are you sure this is correct?\n")
				logString+=("In row " + str(i) + " the total cost is - " + str(totalCost) + ". Are you sure this is correct?\n")
				warningFlag = 1
				
		#Check the placement name column
		#Check it to match the special characer regex above
		plcName = sheet.cell(row=i,column=2).value
		if plcName!=None and plcName!='':
			if not placementPattern.match(str(plcName)):
				print("In row " + str(i) + " the placement name - " + plcName + " is not valid, please revise it!")
				logFile.write("In row " + str(i) + " the placement name - " + plcName + " is not valid, please revise it!\n")
				logString+=("In row " + str(i) + " the placement name - " + plcName + " is not valid, please revise it!\n")
				errorFlag = 1
				

			
	#Check the placement name column on the Campaign Spreadsheet
	#Check to see if there are duplicate placements
	#Go through each line on the CS sheet
	for k in range (2,cscount,1):
		csPlcName = campaignSheetData.cell(row=k,column=4).value
		if csPlcName!=None and csPlcName!='':
			csPlcList.append(csPlcName)	
	if len(csPlcList) != len(set(csPlcList)):
		print("There are duplicate placements - please check the Campaign Spreadsheet tab for the highlighted rows")
		logFile.write("There are duplicate placements - please check the Campaign Spreadsheet tab for the highlighted rows\n")
		logString+=("There are duplicate placements - please check the Campaign Spreadsheet tab for the highlighted rows\n")
		errorFlag = 1
			

	logFile.close()
	return errorFlag, warningFlag, logString
	

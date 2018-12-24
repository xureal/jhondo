def gvalidate(filename, filedir, creds):
	import openpyxl
	import gspread
	from oauth2client.service_account import ServiceAccountCredentials
	import re

	scope = ['https://spreadsheets.google.com/feeds']
	credentials = ServiceAccountCredentials.from_json_keyfile_name(creds, scope)
	gc = gspread.authorize(credentials)


	#Excel file
	#wb = openpyxl.load_workbook(r'D:\Users\andptasi\Desktop\Generic Template.xlsm')
	wb = openpyxl.load_workbook(filename)
	#wbData = openpyxl.load_workbook(r'D:\Users\andptasi\Desktop\Generic Template.xlsm', data_only=True)
	wbData = openpyxl.load_workbook(filename, data_only=True)
	sheet = wb['Media Plan']
	sheetData = wbData['Media Plan']
	campaignSheetData = wbData['Campaign Spreadsheet']
	count = sheet.max_row
	cscount = campaignSheetData.max_row



	#Google Sheets
	sitelistFile = gc.open_by_key('1AuVnBS6M0W9LXmq5m2sJ7-po04zyLQBV3MjbOAJ_rlk')
	sitelistSheet = sitelistFile.get_worksheet(7)

	#Collect list of publishers from the google sheet and put it into a set
	setOfPublishers = set(sitelistSheet.col_values(1))

	#Collect list of languages from the google sheet and put it into a set
	setOfLanguages = set(sitelistSheet.col_values(5))

	#Collect list of countries from the google sheet and put it into a set
	setOfCountries = set(sitelistSheet.col_values(4))

	#Collect list of creative types from the google sheet and put it into a set
	setOfTactic = set(sitelistSheet.col_values(7))

	#Collect list of devices from the google sheet and put it into a set
	setOfDevice = set(sitelistSheet.col_values(8))

	#Collect list of cost model from the google sheet and put it into a set
	setOfCostModel = set(sitelistSheet.col_values(9))

	#Collect list of cost structure from the google sheet and put it into a set
	setOfPlacementType = set(sitelistSheet.col_values(15))
	
	#Collect list of tag type from the google sheet and put it into a set
	setOfTagType = set(sitelistSheet.col_values(6))
	
	#Collect list of ad server from the google sheet and put it into a set
	#setOfAdServer = set(sitelistSheet.col_values(15))
	
	#Collect list of objective from the google sheet and put it into a set
	setOfFormatDetail = set(sitelistSheet.col_values(3))
	
	#Collect list of format from the google sheet and put it into a set
	setOfFormat = set(sitelistSheet.col_values(2))
	
	#Collect list of channel from the google sheet and put it into a set
	setOfChannel = set(sitelistSheet.col_values(6))
	
	
	
	
	#column numbers defined here
	pub_col = 1
	market_col = 2
	lang_col = 3
	tagtype_col = 4
	device_col = 5
	size_col = 6
	format_col = 7
	formatdet_col = 8
	tactic_col = 9
	placement_type_col = 10
	creative1_col = 11
	url1_col = 12
	creative2_col = 13
	url2_col = 14
	creative3_col = 15
	url3_col = 16
	sdate_col = 17
	edate_col = 18
	cost_col = 19
	units_col = 20
	rate_col = 21
	



	#Regex
	sizePattern = re.compile("^[0-9]{1,4}(x)[0-9]{1,4}$")
	urlPattern = re.compile("^(https?|ftp):\/\/(-\.)?([^\s\/?\.#-]+\.?)+(\/[^\s]*)?$")
	datePattern = re.compile("^(0[1-9]|1[0-2])\/(0[1-9]|1\d|2\d|3[01])\/[2][0][1-2][0-9]$")
	placementPattern = re.compile("^[a-zA-Z0-9!@#$()\\-.+ ]*$")
	creativePattern = re.compile("^[a-zA-z0-9]{2,}(_)[0-9]{1,4}(x)[0-9]{1,4}(_)[a-zA-z]{2,3}$")

	logFile = open(filedir+"\\TS_results.txt", 'w')
	logString = ''
	csPlcList = []
	errorFlag = 0
	warningFlag = 0


	#Cycle through the campaign details
	for i in range (3,11,1):
		cmpDetails = str(sheet.cell(row=i, column=2).value)
		if cmpDetails!=None and cmpDetails!='':
			if cmpDetails.strip() != cmpDetails:
				print("The " + str(sheet.cell(row=i,column=1).value) + " field has leading or trailing spaces!")
				logFile.write("The " + str(sheet.cell(row=i,column=1).value) + " field has leading or trailing spaces!\n")
				logString+=("The " + str(sheet.cell(row=i,column=1).value) + " field has leading or trailing spaces!\n")
				errorFlag = 1
				#Cycle through all of the rows on the excel file
	for i in range (12,count,1):
		try:
			# #Check size column to see if the row is a package
			# size = sheet.cell(row=i,column=9).value
			# if size == 'Package' or size == 'package':
				# packagePublisher = sheet.cell(row=i,column=1).value
				# nextRowPublisher = sheet.cell(row=i+1,column=1).value
				# if packagePublisher != nextRowPublisher:
					# print "In row " + str(i) + " the package publisher doesn't match the placement publisher"
					# logFile.write("In row " + str(i) + " the package publisher doesn't match the placement publisher\n")
					# logString+=("In row " + str(i) + " the package publisher doesn't match the placement publisher\n")
					# errorFlag = 1
				# packageDescription = sheet.cell(row=i,column=2).value
				# if not placementPattern.match(str(packageDescription)) or packageDescription == '' or packageDescription == None:
					# print "In row " + str(i) + " the package description is not valid"
					# logFile.write("In row " + str(i) + " the package description is not valid\n")
					# logString+=("In row " + str(i) + " the package description is not valid\n")
					# errorFlag = 1
				# packageMarket = sheet.cell(row=i,column=6).value
				# if packageMarket not in setOfCountries or packageMarket == '' or packageMarket == None:
					# print "In row " + str(i) + " the package market is not valid"
					# logFile.write("In row " + str(i) + " the package market is not valid\n")
					# logString+=("In row " + str(i) + " the package market is not valid\n")
					# errorFlag = 1
				# package_sDate = sheet.cell(row=i,column=18).value
				# package_eDate = sheet.cell(row=i,column=19).value
				# if package_sDate!=None and package_sDate!='':
					# try:
						# pstartDate = package_sDate.strftime('%m/%d/%Y')
					# except:
						# pstartDate = ''
					# if not datePattern.match(pstartDate):
						# print "In row " + str(i) + " the package start date is not valid"
						# logFile.write("In row " + str(i) + " the package start date is not valid\n")
						# logString+=("In row " + str(i) + " the package start date is not valid\n")
						# errorFlag = 1
				# if package_eDate!=None and package_eDate!='':
					# try:
						# pendDate = package_eDate.strftime('%m/%d/%Y')
					# except:
						# pendDate = ''
					# if not datePattern.match(pendDate):
						# print "In row " + str(i) + " the package end date is not valid"
						# logFile.write("In row " + str(i) + " the package end date is not valid\n")
						# logString+=("In row " + str(i) + " the package end date is not valid\n")
						# errorFlag = 1
				# packageCostModel = sheet.cell(row=i,column=20).value
				# if packageCostModel not in setOfCostModel or packageCostModel == '' or packageCostModel == None:
					# print "In row " + str(i) + " the package cost model is not valid"
					# logFile.write("In row " + str(i) + " the package cost model is not valid\n")
					# logString+=("In row " + str(i) + " the package cost model is not valid\n")
					# errorFlag = 1
				# packageUnits = sheet.cell(row=i,column=22).value
				# packageRate = sheet.cell(row=i,column=23).value
				# if packageUnits == '' or packageUnits == None:
					# print "In row " + str(i) + " the package units are blank"
					# logFile.write("In row " + str(i) + " the package units are blank\n")
					# logString+=("In row " + str(i) + " the package units are blank\n")
					# errorFlag = 1
				# if packageRate == '' or packageRate == None:
					# print "In row " + str(i) + " the package rate is blank"
					# logFile.write("In row " + str(i) + " the package rate is blank\n")
					# logString+=("In row " + str(i) + " the package rate is blank\n")
					# errorFlag = 1

			#Cycle through all of the columns on the excel file
			for j in range (1,18,1):
				# if size == 'Package' or size == 'package':
					# if not(sheet.cell(row=i,column=j).value==None or sheet.cell(row=i,column=j).value=='') and (j!=1 and j!=2 and j!=6 and j!=9 and j!=18 and j!=19 and j!=20 and j!=22 and j!=23 and j!=24 and j!=25):
						# print "In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " must be blank if the line is a package"
						# logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " must be blank if the line is a package\n")
						# logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " must be blank if the line is a package\n")
						# errorFlag = 1

				#Check for leading and trailing spaces
				cellValue = str(sheet.cell(row=i,column=j).value)
				if cellValue!=None and cellValue!='':
					if cellValue.strip() != cellValue:
						print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " has leading or trailing spaces!")
						logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " has leading or trailing spaces!\n")
						logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " has leading or trailing spaces!\n")
						errorFlag = 1

				#Check if publisher is there but other columns are blank
				publisher = sheet.cell(row=i,column=pub_col).value
				size = sheet.cell(row=i,column=size_col).value			
				if (publisher!=None and publisher!='') and not (size=='Package' or size=='package'):
					if sheet.cell(row=i,column=j).value==None or sheet.cell(row=i,column=j).value=='':
						if (j==creative2_col or j==url2_col or j==creative3_col or j==url3_col) and ((sheet.cell(row=i,column=creative1_col).value==None or sheet.cell(row=i,column=creative1_col).value=='') and (sheet.cell(row=i,column=url1_col).value==None or sheet.cell(row=i,column=url1_col).value=='')):
							print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!")
							logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
							logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
							errorFlag = 1
						elif (j!=creative2_col and j!=url2_col and j!=creative3_col and j!=url3_col and j!=units_col and j!=rate_col):
							print("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!")
							logFile.write("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
							logString+=("In row " + str(i) + " the column - " + str(sheet.cell(row=11,column=j).value) + " is blank!\n")
							errorFlag = 1
							
			# Check publisher column
			# Check if it's present in the publisher list on google sheet
			publisher = sheet.cell(row=i,column=pub_col).value
			if publisher!=None and publisher!='':
				if publisher not in setOfPublishers:
					print("In row " + str(i) + " the publisher - " + publisher + " is not in the list of defined publishers!")
					logFile.write("In row " + str(i) + " the publisher - " + publisher + " is not in the list of defined publishers!\n")
					logString+=("In row " + str(i) + " the publisher - " + publisher + " is not in the list of defined publishers!\n")
					errorFlag = 1

			#Check size column
			#Match it to regex above of 1-4 numbers, lowercase x, 1-4 numbers
			size = sheet.cell(row=i,column=size_col).value
			if size!=None and size!='' and size!='Package' and size!='package':
				if not sizePattern.match(size):
					print("In row " + str(i) + " the size - " + size + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the size - " + size + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the size - " + size + " is not valid, please revise it!\n")
					errorFlag = 1
					
			#Check language column
			#Check if it's a 3 letter code present on the google sheet
			language = sheet.cell(row=i,column=lang_col).value
			if language!=None and language!='':
				if language not in setOfLanguages:
					print("In row " + str(i) + " the language - " + language + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the language - " + language + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the language - " + language + " is not valid, please revise it!\n")
					errorFlag = 1
			#Check country column
			#Check if it's a 2 letter ISO country code that's on the google sheet
			country = sheet.cell(row=i,column=market_col).value
			if country!=None and country!='':
				if country not in setOfCountries:
					print("In row " + str(i) + " the country - " + country + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the country - " + country + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the country - " + country + " is not valid, please revise it!\n")
					errorFlag = 1

			#Check creative type column
			#Check if it's present in the creative type list on google sheet
			placementType = sheet.cell(row=i,column=placement_type_col).value
			if placementType!=None and placementType!='':
				if placementType not in setOfPlacementType:
					print("In row " + str(i) + " the placement type - " + placementType + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the placement type - " + placementType + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the placement type - " + placementType + " is not valid, please revise it!\n")
					errorFlag = 1

			#Check device column
			#Check if it's present in the device list on google sheet
			device = sheet.cell(row=i,column=device_col).value
			if device!=None and device!='':
				if device not in setOfDevice:
					print("In row " + str(i) + " the device - " + device + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the device - " + device + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the device - " + device + " is not valid, please revise it!\n")
					errorFlag = 1
					
			#Check cost model column
			#Check if it's present in the cost model list on google sheet
			costModel = sheet.cell(row=i,column=cost_col).value
			if costModel!=None and costModel!='':
				if costModel not in setOfCostModel:
					print("In row " + str(i) + " the cost model - " + costModel + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the cost model - " + costModel + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the cost model - " + costModel + " is not valid, please revise it!\n")
					errorFlag = 1
					
			#Check tag type column
			#Check if it's present in the tag type list on google sheet
			tagType = sheet.cell(row=i,column=tagtype_col).value
			if tagType!=None and tagType!='':
				if tagType not in setOfTagType:
					print("In row " + str(i) + " the tag type - " + tagType + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the tag type - " + tagType + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the tag type - " + tagType + " is not valid, please revise it!\n")
					errorFlag = 1
					
			#Check ad server column
			#Check if it's present in the ad server list on google sheet
			formatDetails = sheet.cell(row=i,column=formatdet_col).value
			if formatDetails!=None and formatDetails!='':
				if formatDetails not in setOfFormatDetail:
					print("In row " + str(i) + " the format details - " + formatDetails + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the format details - " + formatDetails + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the format details - " + formatDetails + " is not valid, please revise it!\n")
					errorFlag = 1
					
			#Check objective column
			#Check if it's present in the objective list on google sheet
			tactic = sheet.cell(row=i,column=tactic_col).value
			if tactic!=None and tactic!='':
				if tactic not in setOfTactic:
					print("In row " + str(i) + " the tactic - " + tactic + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the tactic - " + tactic + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the tactic - " + tactic + " is not valid, please revise it!\n")
					errorFlag = 1
					
			#Check format column
			#Check if it's present in the format list on google sheet
			format = sheet.cell(row=i,column=format_col).value
			if format!=None and format!='':
				if format not in setOfFormat:
					print("In row " + str(i) + " the format - " + format + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the format - " + format + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the format - " + format + " is not valid, please revise it!\n")
					errorFlag = 1
					
			#Check channel column
			#Check if it's present in the channel list on google sheet
			# channel = sheet.cell(row=i,column=5).value
			# if channel!=None and channel!='':
				# if channel not in setOfChannel:
					# print "In row " + str(i) + " the channel - " + channel + " is not valid, please revise it!"
					# logFile.write("In row " + str(i) + " the channel - " + channel + " is not valid, please revise it!\n")
					# logString+=("In row " + str(i) + " the channel - " + channel + " is not valid, please revise it!\n")
					# errorFlag = 1

			# #Check cost structure column
			# #Check if it's present in the cost structure list on google sheet
			# costStructure = sheet.cell(row=i,column=17).value
			# if costStructure!=None and costStructure!='':
				# if costStructure not in setOfCostStructure:
					# print "In row " + str(i) + " the cost structure - " + costStructure + " is not valid, please revise it!"
					# logFile.write("In row " + str(i) + " the cost structure - " + costStructure + " is not valid, please revise it!\n")
					# logString+=("In row " + str(i) + " the cost structure - " + costStructure + " is not valid, please revise it!\n")
					# errorFlag = 1
					
			#Check URL column
			#Match it to the regex above for URL matching
			url1 = sheet.cell(row=i,column=url1_col).value
			url2 = sheet.cell(row=i,column=url2_col).value
			url3 = sheet.cell(row=i,column=url3_col).value
			if url1!=None and url1!='':
				if not urlPattern.match(url1):
					print("In row " + str(i) + " url1 - " + url1 + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " url1 - " + url1 + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " url1 - " + url1 + " is not valid, please revise it!\n")
					errorFlag = 1
			if url2!=None and url2!='':
				if not urlPattern.match(url2):
					print("In row " + str(i) + " url2 - " + url2 + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " url2 - " + url2 + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " url2 - " + url2 + " is not valid, please revise it!\n")
					errorFlag = 1
			if url3!=None and url3!='':
				if not urlPattern.match(url3):
					print("In row " + str(i) + " url3 - " + url3 + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " url3 - " + url3 + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " url3 - " + url3 + " is not valid, please revise it!\n")
					errorFlag = 1
					
			#Check date column
			#Match it to regex above of mm/dd/yyyy
			sDate = sheet.cell(row=i,column=sdate_col).value
			eDate = sheet.cell(row=i,column=edate_col).value


			if sDate!=None and sDate!='':
				try:
					startDate = sDate.strftime('%m/%d/%Y')
				except:
					startDate = ''
				if not datePattern.match(startDate):
					print("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the start date - " + startDate + " is not valid, please revise it!\n")
					errorFlag = 1
			if eDate!=None and eDate!='':
				try:
					endDate = eDate.strftime('%m/%d/%Y')
				except:
					endDate = ''
				if not datePattern.match(endDate):
					print("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!")
					logFile.write("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!\n")
					logString+=("In row " + str(i) + " the end date - " + endDate + " is not valid, please revise it!\n")
					errorFlag = 1

			#Check total USD column
			#Flag a warning if cost is above threshold
			totalCost = sheetData.cell(row=i,column=24).value
			if totalCost !=None and totalCost!='':
				if totalCost > 2000:
					print("In row " + str(i) + " the total cost is - " + str(totalCost) + ". Are you sure this is correct?")
					logFile.write("In row " + str(i) + " the total cost is - " + str(totalCost) + ". Are you sure this is correct?\n")
					logString+=("In row " + str(i) + " the total cost is - " + str(totalCost) + ". Are you sure this is correct?\n")
					warningFlag = 1
					
			#Check the placement name column
			#Check it to match the special characer regex above
			# plcName = sheet.cell(row=i,column=plcdesc_col).value
			# if plcName!=None and plcName!='':
				# if not placementPattern.match(str(plcName)):
					# print "In row " + str(i) + " the placement name - " + plcName + " is not valid, please revise it!"
					# logFile.write("In row " + str(i) + " the placement name - " + plcName + " is not valid, please revise it!\n")
					# logString+=("In row " + str(i) + " the placement name - " + plcName + " is not valid, please revise it!\n")
					# errorFlag = 1
		except Exception as e:
			print(e)
			print("In row " + str(i) + " there is a major error - please make sure all characters are valid")
			logFile.write("In row " + str(i) + " there is a major error - please make sure all characters are valid\n")
			logString+=("In row " + str(i) + " there is a major error - please make sure all characters are valid\n")
			errorFlag = 1
			continue


		#Check the creative name column
		#Check if it matches the naming convention (client_creative_size_language)
		# crtName1 = sheet.cell(row=i,column=12).value
		# crtName2 = sheet.cell(row=i,column=14).value
		# crtName3 = sheet.cell(row=i,column=16).value
		# if crtName1!=None and crtName1!='' and crtName1.lower()!='tracking':
			# if not creativePattern.match(crtName1):
				# print "In row " + str(i) + " the creative name - " + crtName1 + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the creative name - " + crtName1 + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the creative name - " + crtName1 + " is not valid, please revise it!\n")
				# errorFlag = 1
		# if crtName2!=None and crtName2!='' and crtName2.lower()!='tracking':
			# if not creativePattern.match(crtName2):
				# print "In row " + str(i) + " the creative name - " + crtName2 + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the creative name - " + crtName2 + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the creative name - " + crtName2 + " is not valid, please revise it!\n")
				# errorFlag = 1
		# if crtName3!=None and crtName3!='' and crtName3.lower()!='tracking':
			# if not creativePattern.match(crtName3):
				# print "In row " + str(i) + " the creative name - " + crtName3 + " is not valid, please revise it!"
				# logFile.write("In row " + str(i) + " the creative name - " + crtName3 + " is not valid, please revise it!\n")
				# logString+=("In row " + str(i) + " the creative name - " + crtName3 + " is not valid, please revise it!\n")
				# errorFlag = 1
			
	#Check the placement name column on the Campaign Spreadsheet
	#Check to see if there are duplicate placements
	#Go through each line on the CS sheet
	for k in range (2,cscount,1):
		csPlcName = campaignSheetData.cell(row=k,column=4).value
		if csPlcName!=None and csPlcName!='':
			csPlcList.append(csPlcName)	
	if len(csPlcList) != len(set(csPlcList)):
		print("There are duplicate placements - please check the Campaign Spreadsheet tab for the highlighted rows")
		logFile.write("There are duplicate placements - please check the Campaign Spreadsheet tab for the highlighted rows\n")
		logString+=("There are duplicate placements - please check the Campaign Spreadsheet tab for the highlighted rows\n")
		errorFlag = 1
			

	logFile.close()
	return errorFlag, warningFlag, logString
	

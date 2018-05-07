import argparse
from oauth2client import file, client, tools
from apiclient.discovery import build
from httplib2 import Http
import config
import dictionaries
import openpyxl
import sys
import os
import xlrd
from datetime import datetime

############stupid shit to get started##################

parser = argparse.ArgumentParser(parents=[tools.argparser])
flags = parser.parse_args()

CLIENT_SECRET = 'client_secret.json' # downloaded JSON file
SCOPES = ('https://www.googleapis.com/auth/dfareporting','https://www.googleapis.com/auth/dfatrafficking')

store = file.Storage('storage.json')
creds = store.get()
if not creds or creds.invalid:
    flow = client.flow_from_clientsecrets(CLIENT_SECRET, SCOPES)
    creds = tools.run_flow(flow, store, flags)

DFA = build('dfareporting', 'v3.0', http=creds.authorize(Http()))

print('Connected to Google servers')

##############end stupid shit##########################

#Input the values for advertiser name and campaign ID
advertiser_name = input('Please enter the advertiser name: ')
campaign_id = input('Please enter the campaign ID: ')
print('\n')

#Check if the entered advertiser is in the alias dict of the config file or the main advertiser dict, if it's in neither then ask user to input the advertiser ID or choose to exit the program
if advertiser_name in config.alias:
    advertiser_id = dictionaries.advertisers(config.alias[advertiser_name])[0]
    profile_id = dictionaries.advertisers(config.alias[advertiser_name])[2]
elif dictionaries.advertisers(advertiser_name) != 0:
    advertiser_id = dictionaries.advertisers(advertiser_name)[0]
    profile_id = dictionaries.advertisers(advertiser_name)[2]
else:
    print('Advertiser doesn\'t exist, please add the advertiser to the dictionary to proceed')
    print('Terminating...')
    sys.exit()

#Initizlize all of the required variables
workbook = openpyxl.load_workbook(config.feed_path)
sheet = workbook.get_sheet_by_name(config.feed_sheet)
row_count = 0
duplicate_count = 0
dcm_placements_name = []
duplicate_placements = []
new_placements = []
sdates_string = []
sdates_date = []
edates_string = []
edates_date = []

fetch_dcm_campaign = DFA.campaigns().get(profileId=profile_id,id=campaign_id).execute() #Fetch the campaign from DCM by ID
fetch_dcm_placement = DFA.placements().list(profileId=profile_id,campaignIds=campaign_id,advertiserIds=advertiser_id,archived=False).execute() #Fetch the placement from the DCM campaign by name


#Go through the temp directory and find the upload sheet, then go through all of the start and end date and add to list
for root, dirs, files in os.walk(config.temp_folder, topdown=True):
	if len(files)>0:
		for name in files:
			if name.endswith('.xls'):
				filepath = root+'\\'+name
				filename = name
				wb = xlrd.open_workbook(filename=filepath)
				upload_sheet = wb.sheet_by_name('Campaign Spreadsheet')
				for i in range(1, upload_sheet.nrows, 1):
					sdate_cell_value = upload_sheet.cell(rowx=i,colx=11).value
					if sdate_cell_value != '':
						sdates_string.append(sdate_cell_value)					
					edate_cell_value = upload_sheet.cell(rowx=i,colx=12).value
					if edate_cell_value != '':
						edates_string.append(edate_cell_value)

#Change the dates from string to date format that is required in DCM
for j in range(0,len(sdates_string),1):
	sdate_object = datetime.strptime(sdates_string[j], '%m/%d/%y').date()
	sdates_date.append(sdate_object)
for k in range(0,len(sdates_string),1):
	edate_object = datetime.strptime(edates_string[k], '%m/%d/%y').date()
	edates_date.append(edate_object)


if str(min(sdates_date)) < fetch_dcm_campaign['startDate']:
    print('The upload sheet has placements that have a start date earlier than the campaign start date')
elif str(max(edates_date)) > fetch_dcm_campaign['endDate']:
    print('The upload sheet has placements that have an end date later than the campaign end date')
else:
    print('There are no issues with the dates of the placements on the upload sheet')

print('\n')


for p in range(0,len(fetch_dcm_placement['placements']),1):
    dcm_placements_name.append(fetch_dcm_placement['placements'][p]['name'])

for i in range(1,sheet.max_row,1): #Find the total number of non-empty rows in the testtest sheet
    if sheet['A'+str(i)].value!=None and sheet['A'+str(i)].value!='':
        row_count += 1

for j in range(2,row_count+1,1):
    feed_placement_name = sheet['A'+str(j)].value #Get placement name from the feed

    if feed_placement_name in dcm_placements_name: #Compare the placement names on the testtest sheet with the ones in the DCM campaign
        duplicate_placements.append(feed_placement_name)
        duplicate_count += 1 #If placement is found in the DCM campaign increment the duplicate count
    else:
        new_placements.append(feed_placement_name)

if duplicate_count > 0: #Display the below is there are any duplicate placements
    print('Some of the placements in the feed already exist in DCM, enter y to print the placements names or press enter to exit')
    print('Total duplicate placements - ' + str(duplicate_count))
    print('Existing placements / Total placements on the sheet - ' + str(duplicate_count) + '/' + str(row_count-1))
    duplicate_display = input('Display duplicate placements?: ')
    print('\n')
    if duplicate_display != '':
        for k in range(0,len(duplicate_placements),1):
            print(duplicate_placements[k])
        print('__NEW PLACEMENTS BELOW__')
        for kk in range(0,len(new_placements),1):
            print(new_placements[kk])
    else:
        print('Terminating...')
        sys.exit()
else:
    print('None of the placements on the feed have been found in the DCM campaign')
